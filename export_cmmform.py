from flask import *
from flask_cors import CORS
import pyodbc
import pandas as pd
import openpyxl
from openpyxl import load_workbook,styles
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Fill, Border, Side
from openpyxl.styles import Font
import traceback
import io
app = Flask(__name__)
CORS(app)
@app.route("/CMM/export", methods=['POST', 'GET'])
def exportCMM():
        payload = request.get_json()
        formname = payload["formName"]
        dmclist1 = payload.get('dmc')   
        # dmclist = list(set(dmclist))
        dmclist = []
        for item in dmclist1:
            if item not in dmclist:
                dmclist.append(item)
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)   
        filename = pd.read_sql("select top 1 filename from cmmformdata where Formname ='"+formname+"'",conn).loc[0,"filename"]
        wb_sample = load_workbook('cmmform/'+formname+'/'+filename)
        sheet=wb_sample.active
        my_red = openpyxl.styles.colors.Color(rgb='00ff5555')
        my_gray = openpyxl.styles.colors.Color(rgb='00d9d9d9')
        daycl = openpyxl.styles.colors.Color(rgb='0055ff55')
        nightcl = openpyxl.styles.colors.Color(rgb='00FFA500')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_fill2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_gray)
        dayfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=daycl)
        nightfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=nightcl)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border1 = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        
        if len(dmclist) > 3:
            try:
                sheet.unmerge_cells(start_row=21,end_row=22,start_column=19,end_column=19)
            except:
                pass
            sheet.insert_cols(17,len(dmclist)-3)
            for i in range(len(dmclist)-3):
                
                for row_num in range(23, sheet.max_row + 1):
                    sheet.cell(row_num,17+i).border = border1
                    sheet.cell(row_num,17+i).alignment = Alignment(horizontal='center',vertical='center')
                    sheet.cell(row_num,17+i).font = Font(name="Arial")
                for row in range(2,10):
                    sheet.cell(row,17+i).border = Border(left=Side(style='thin'), right=Side(style='dashed'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in range(11,21):
                    sheet.cell(row,17+i).border = Border(top=Side(style='dashed'))
                sheet.cell(10,17+i).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                sheet.cell(21,17+i).border = border
                sheet.cell(22,17+i).border = border
                sheet.cell(23,17+i).border = border
                sheet.cell(21,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(22,17+i).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(23,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(21,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).font = Font(name="Arial")
                sheet.cell(23,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).fill = my_fill2
            # sheet.merge_cells(start_row=21,end_row=22,start_column=15+len(dmclist),end_column=15+len(dmclist))
    
        for idmc in range(len(dmclist)):
            try:
                # print("SELECT [DMC],[Line],[TimeSave],[id],[actual],[Result],[CodePurpose] FROM [QC].[dbo].[CMMdata] where dmc = '"+dmclist[idmc]+"'")
                idlist = pd.read_sql("select link as idx from cmmformdata where formname = '"+formname+"' order by Position",conn)
                data = pd.read_sql("SELECT [DMC],[Line],[TimeSave],[id],[actual],[Result],[CodePurpose] FROM [QC].[dbo].[CMMdata] where dmc = N'"+dmclist[idmc]+"'",conn)
                print(idmc,len(dmclist))
                if data.loc[0,"TimeSave"].hour < 8 or data.loc[0,"TimeSave"].hour > 19:
                    sheet.cell(21,idmc+16).fill = nightfill
                else:
                    sheet.cell(21,idmc+16).fill = dayfill
                sheet.cell(21,idmc+16).value = str(data.loc[0,"TimeSave"])[:-7]
                sheet.cell(22,idmc+16).value = data.loc[0,"Line"] +' '+ dmclist[idmc]
                print(data.loc[0,"Line"] +' '+ dmclist[idmc])
                sheet.cell(23,idmc+16).value = data.loc[0,"CodePurpose"]
                sheet.cell(23,idmc+16).border = border
                data.set_index("id", inplace=True)
                for i in range(len(idlist)):
                    try:
                        sheet.cell(24+i,idmc+16).value = round(float(data.loc[idlist.loc[i,"idx"],"actual"]),3)
                        sheet.cell(24+i,idmc+16).number_format = '0.000'
                        if "No Tol" not in idlist.loc[i,"idx"]:                            
                            if data.loc[idlist.loc[i,"idx"],"Result"].strip() == "":
                                if data.loc[idlist.loc[i,"idx"].replace(".XA",'').replace(".YA",'').replace(".ZA",'').replace(".R",'').replace(".PH",''),"Result"].strip() != "OK":
                                    sheet.cell(24+i,idmc+16).fill = my_fill
                            elif data.loc[idlist.loc[i,"idx"],"Result"].strip() != "OK":
                                sheet.cell(24+i,idmc+16).fill = my_fill
                    except Exception as e:
                        sheet.cell(24+i,idmc+16).fill = my_fill
                    
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"nominal"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"lowertol"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"uppertol"]
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"unit"]
                        # sheet.cell(3+i,idmc+15).value = "CMM Machine"
                        # sheet.cell(3+i,idmc+15).value = data.loc[i,"actual"]
                
            except:
                print(dmclist[idmc])
        # data = data.to_json()
        output = io.BytesIO()

        # add headers
        wb_sample.save(output)
        # writer.save()
        output.seek(0)

        return Response(output, mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=CMM_Data.xls"})

app.run(host='0.0.0.0',port=12345,debug=True)
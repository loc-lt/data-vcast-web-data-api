from os import access
from constants.http_status_code import HTTP_200_OK, HTTP_201_CREATED, HTTP_400_BAD_REQUEST, HTTP_401_UNAUTHORIZED, HTTP_409_CONFLICT
from flask import Blueprint, app, request, jsonify
from werkzeug.security import check_password_hash, generate_password_hash
from flask_jwt_extended import jwt_required, create_access_token, create_refresh_token, get_jwt_identity
from flasgger import swag_from
from database import *
from ERROR import *
import traceback
import pandas as pd
import requests
import  os
import shutil
from openpyxl import load_workbook
from auth_middleware import *
import json

CMM = Blueprint("CMM", __name__, url_prefix="/api/v1/CMM")

# Viết lại các api cho A. Hiển
# Lấy danh sách tất cả product
@CMM.get('/products') 
@swag_from('./docs/CMM/get_products.yaml')
@token_required
def get_products():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("select distinct(Product) as Product from CMMData order by Product")
        product_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if product_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(product_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in product_list:
            file_name = item
            ret["data"].append(file_name[0])
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "products").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/machines') 
@swag_from('./docs/CMM/get_machines.yaml')
@token_required
def get_machines():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("select distinct(CMMmachine) as Product from CMMData order by CMMmachine")
        machine_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if machine_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(machine_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in machine_list:
            file_name = item
            ret["data"].append(file_name[0])
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "machines").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/forms') 
@swag_from('./docs/CMM/get_forms.yaml')
@token_required
def get_forms():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("select distinct(FormName) as Fname from CMMFormData")
        form_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if form_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(form_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in form_list:
            file_name = item
            ret["data"].append(file_name[0])
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "forms").append_new_line()
        return jsonify(ret),500
    
@CMM.post('') 
@swag_from('./docs/CMM/data.yaml')
@token_required_and_permissions
def get_data(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 7
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Ngày
        day_start = request.json['dayStart']
        day_end = request.json['dayEnd']

        # Giờ
        time_start = request.json['timeStart']
        time_end = request.json['timeEnd']

        # Ngày giờ
        datetime_start = day_start + " " + time_start
        datetime_end = day_end + " " + time_end
        
        product = request.json["product"]
        machine = request.json["machine"]
        dmc = request.json["dmc"]

        print("SELECT Idx, TimeStart, TimeSave, requester, CodePurpose, DMC, Line, CMMmachine,CMMCode, Operator, (CASE WHEN (SELECT count(value) FROM STRING_SPLIT(Product, ' ')) >= 2 THEN (SELECT dbo.ConcatValuesInColumn(Product, 2)) ELSE Product END) as Product, SUM(CASE WHEN Result != '' THEN 1 ELSE 0 END) AS Total,SUM(CASE WHEN id like '%No Tol%' THEN 1 ELSE 0 END) AS NoTol, SUM(CASE WHEN id like '%In Tol%' THEN 1 ELSE 0 END) AS InTol,SUM(CASE WHEN Result = 'OK'  and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS OK_Total,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS NG_Total,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%' or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS OK_C,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%'or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS NG_C,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS OK_CM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS NG_CM,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS OK_MM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS NG_MM FROM CMMdata where timesave >'"+datetime_start+"' and timesave <'"+datetime_end+"' and dmc like '%"+dmc+"%' and product like '%"+product+"%' and CMMmachine like '%"+machine+"%' GROUP BY Idx,DMC,TimeStart,requester, CodePurpose,CMMCode, Line, CMMmachine, Operator, TimeSave, Product")
        cursor.execute("SELECT Idx, TimeStart, TimeSave, requester, CodePurpose, DMC, Line, CMMmachine,CMMCode, Operator, (CASE WHEN (SELECT count(value) FROM STRING_SPLIT(Product, ' ')) >= 2 THEN (SELECT dbo.ConcatValuesInColumn(Product, 2)) ELSE Product END) as Product, SUM(CASE WHEN Result != '' THEN 1 ELSE 0 END) AS Total,SUM(CASE WHEN id like '%No Tol%' THEN 1 ELSE 0 END) AS NoTol, SUM(CASE WHEN id like '%In Tol%' THEN 1 ELSE 0 END) AS InTol,SUM(CASE WHEN Result = 'OK'  and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS OK_Total,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' THEN 1 ELSE 0 END) AS NG_Total,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%' or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS OK_C,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and (id like '% C\_%' ESCAPE '\\' or id like '% C.%'or id like '% C %' or id like '% CC %') THEN 1 ELSE 0 END) AS NG_C,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS OK_CM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id like '%CM%' THEN 1 ELSE 0 END) AS NG_CM,SUM(CASE WHEN Result = 'OK' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS OK_MM,SUM(CASE WHEN Result = 'NG' and id not like '%No Tol%' and id not like '%In Tol%' and id not like '%CM%' and id not like '% C\_%' ESCAPE '\\' and id not like '% C.%' and id not like '% C %' and id not like '% CC %' THEN 1 ELSE 0 END) AS NG_MM FROM CMMdata where timesave >'"+datetime_start+"' and timesave <'"+datetime_end+"' and dmc like '%"+dmc+"%' and product like '%"+product+"%' and CMMmachine like '%"+machine+"%' GROUP BY Idx, DMC,TimeStart,requester, CodePurpose,CMMCode, Line, CMMmachine, Operator, TimeSave, Product")
        form_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if form_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(form_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)

        for item in form_list:
            id, time_start, time_save, requester, code_purpose, dmc, line, cmm_machine, cmm_code, operator, product_temp, total, notol, intol, ok_total, ng_total, cc_ok, cc_ng, cm_ok, cm_ng, mm_ok, mm_ng = item
            ret["data"].append({
                "id": id,
                "timeStart": str(time_start),
                "timeSave": str(time_save),
                "requester": requester,
                "codePurpose": code_purpose,
                "dmc": dmc,
                "line": line,
                "product": product_temp,
                "cmmMachine": cmm_machine,
                "cmmCode": cmm_code,
                "operator": operator,
                "total": total,
                "noTol": notol,
                "inTol": intol,
                "ccOK": cc_ok,
                "ccNG": cc_ng,
                "cmOK": cm_ok,
                "cmNG": cm_ng,
                "mmOK": mm_ok,
                "mmNG": mm_ng
            })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "cmm_data").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/chart') 
@swag_from('./docs/CMM/chart_data.yaml')
@token_required_and_permissions
def chart_data(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 12
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Ngày, tên máy
        time_start = request.json['timeStart']
        time_end = request.json['timeEnd']

        cmm_machine = str(request.json["cmmMachine"]).replace('[','(').replace(']',')')

        print(time_start)
        print(time_end)
        print(cmm_machine)

        print("select timestart, timesave, cmmmachine from [QC].[dbo].[CMMdata] where cmmmachine in "+cmm_machine+" and timesave > '"+time_start+"' and timestart < '"+time_end+"' group by  timestart, timesave, cmmmachine order by cmmmachine,timesave")
        cursor.execute("select timestart, timesave, cmmmachine from [QC].[dbo].[CMMdata] where cmmmachine in "+cmm_machine+" and timesave > '"+time_start+"' and timestart < '"+time_end+"' group by  timestart, timesave, cmmmachine order by cmmmachine,timesave")
        data_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if data_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data':[]
            }
        
        if len(data_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)

        for item in data_list:
            time_start, time_save, cmm_machine = item
            ret["data"].append({
                    "x": cmm_machine,
                    "y": [int(time_start.timestamp()*1000+25200000), int(time_save.timestamp()*1000+25200000)]
                })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "cmm_chart_data").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/sixpack_products') 
@swag_from('./docs/CMM/sixpack_products.yaml')
@token_required
def get_sixpack_products():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("select distinct(Product) as Product from CMMsixpackform  order by Product")
        sixpack_products = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if sixpack_products == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(sixpack_products) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in sixpack_products:
            product_name = item
            ret["data"].append(product_name[0])
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "machines").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/sixpack_forms/<string:dmc>') 
@swag_from('./docs/CMM/get_sixpack_forms.yaml')
@token_required_and_permissions
def get_sixpack_form(role, permissions, dmc):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 13
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401
        
        # Truy cập Database để lấy dữ liệu
        cursor.execute("select * FROM [QC].[dbo].[CMMsixpackform] where product = '"+dmc+"' order by name")
        all_data = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if all_data == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(all_data) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in all_data: 
            product, cmm, name = item
            ret["data"].append({
                "cmm": cmm,
                "name": name
            })  

        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "sixpack_forms").append_new_line()
        return jsonify(ret), 500
    
@CMM.get('/data/<string:dmc>/<string:datetime>')
@swag_from('./docs/CMM/data_details.yaml')
@token_required
def data_details(dmc, datetime):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("SELECT IDx, TimeSave, Actual, Nominal, Uppertol, Lowertol, Unit, Result, Id FROM [QC].[dbo].[CMMdata] where DMC = '"+dmc+"' and TimeSave = '"+datetime+"'")
        cmm_datas = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if cmm_datas == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data':[]
            }
        
        if len(cmm_datas) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in cmm_datas:
            idx, time_save, actual, nominal, uppertol, lowertol, unit, result, id = item
            ret["data"] += [{
            "index":idx,
            "timeSave":time_save,
            "actual":actual,
            "nominal":nominal,
            "upper":uppertol,
            "lower":lowertol,
            "unit":unit,
            "result":result,
            "name":id
        }]   
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "data_details").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/form_manager') 
@swag_from('./docs/CMM/form_manager.yaml')
@token_required
def get_form_manager():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("select distinct(FormName), FileName from CMMFormdata group by FormName, FileName")
        form_manager_list = cursor.fetchall()

        print(len(form_manager_list))

        # Nếu lấy dữ liệu ra trống
        if form_manager_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(form_manager_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in form_manager_list:
            form_name, file_name = item
            ret["data"].append({
                'formName': form_name,
                'fileName': file_name
            })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "form_manager").append_new_line()
        return jsonify(ret),500

@CMM.post('/import_data') 
@swag_from('./docs/CMM/import_data.yaml')
@token_required
def test_import_file():
    print("Vô nè")
    try:
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234; Database=KnifeCNCSystem; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Kiểm tra xem 'data' có nằm trong list files không
        if 'data' not in request.files:
            ret = {
                'status':False,
                'message':'No file part'
            }
            return jsonify(ret),400

        # Lấy file storage object
        uploaded_file = request.files['data']

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                "data": []
            }
        
        # Lấy file_name
        file_name = uploaded_file.filename

        # Lấy tên máy
        if len(file_name.split('OP')) == 2:
            machine_name = file_name.split('OP')[1].split('-')[1].split('.')[0]
        else:
            machine_name = file_name.split('OP')[2].split('-')[1].split('.')[0]

        # Xóa toàn bộ dữ liệu bảng Bom với Machine_Name là machine_name phía trên
        cursor.execute("delete from Bom where Machine_Name = '"+machine_name+"'")
        conn.commit()

        # Lấy danh sách các sheet
        xls_sheetnames = pd.ExcelFile(uploaded_file).sheet_names

        # Duyệt qua danh sách các sheetname rồi lấy dữ liệu của từng sheetname
        for sheetname in xls_sheetnames:
            sheet_data = pd.read_excel(uploaded_file, sheet_name=sheetname)
            
            # lấy op
            op = sheetname
            
            print(sheet_data.head())

            # lấy dmc_product
            dmc_product_text = sheet_data.at[2, "TIÊU CHUẨN DAO CỤ\nTHE STANDARD OF TOOL USEAGE"]
            dmc_product = dmc_product_text.split(":")[1].strip()[:7]

            # duyệt qua từng dòng của dataframe để lấy dữ liệu
            for idx, row in sheet_data.iterrows():
                if isinstance(row[0], int):
                    print(row)
                    print("row[0]", row[0])
                    
                    # Lấy tool_holder
                    tool_holder = row[1]

                    if tool_holder is None:
                        break
                    elif tool_holder == 'N/A':
                        tool_holder = '0'
                    else:
                        tool_holder = tool_holder.split('T')[1]
                        if '(' in tool_holder:
                            tool_holder = tool_holder.split('(')[0]

                    # Lấy tool_type và arbor_type
                    tool_type = row[3]
                    print("arbor_type", row[7])
                    arbor_type = row[7].replace("\n", "").strip()

                    # Lấy tool_useage
                    tool_useage_to_check = row[12]
                    
                    if isinstance(tool_useage_to_check, int):
                        tool_useage = tool_useage_to_check
                    else:
                        if len(tool_useage_to_check.split(":")) > 1:
                            if len(tool_useage_to_check.split(":")) == 2:
                                if ')' in tool_useage_to_check.split(":")[1]:
                                    tool_useage = tool_useage_to_check.split(":")[1].replace(')', '')
                                else:
                                    tool_useage = tool_useage_to_check.split(":")[1]
                            elif len(tool_useage_to_check.split(":")) == 3:
                                if ":" in tool_useage_to_check.split()[2]:
                                    tool_useage = tool_useage_to_check.split()[2][1:]
                                else:
                                    tool_useage = tool_useage_to_check.split()[2]
                            else:
                                if 'PCS' in tool_useage_to_check.split(":")[1].split("\n")[0].strip():
                                    tool_useage = tool_useage_to_check.split(":")[1].split("\n")[0].strip().replace(' PCS', '')
                                else:
                                    tool_useage = tool_useage_to_check.split(":")[1].split("\n")[0].strip()
                        else:
                            tool_useage = 0

                    # Lấy diameter và tolerance
                    diameter_tolerance = row[15]
                    diameter_tolerance = diameter_tolerance.split('+')

                    diameter = diameter_tolerance[0]
                    tolerance = diameter_tolerance[1]
                
                    print((machine_name, dmc_product, op, tool_type, tool_holder, arbor_type, tool_useage, diameter, tolerance))
                
                    cursor.execute( 
                            '''INSERT INTO Bom VALUES (?,?,?,?,?,?,?,?,?) ''',
                                machine_name,
                                dmc_product,
                                op,
                                tool_type, 
                                tool_holder,
                                arbor_type, 
                                tool_useage,
                                diameter,
                                tolerance
                        )
                    cursor.commit()

        cursor.execute("select * FROM [KnifeCNCSystem].[dbo].[Bom]")
        all_records = cursor.fetchall()
        conn.close()
        
        for item in all_records:
            idx, machine_name, dmc_product, op, tool_type, tool_holder, arbor_type, tool_useage, diameter, tolerance = item
            ret['data'].append({
                'machine_name':machine_name.strip(),
                'dmc_product':dmc_product.strip(),
                'op':op.strip(),
                'tool_type':tool_type.strip(),
                'tool_holder':tool_holder,
                'arbor_type':arbor_type.strip(),
                'tool_usage':tool_useage,
                'diameter':diameter,
                'tolerance':tolerance
            })
            
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "form_manager").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/form_manager/<string:form_name>') 
@swag_from('./docs/CMM/form_manager_detail.yaml')
@token_required
def get_form_manager_detail(form_name):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("SELECT Position, CircleNum, Characteristic, Link FROM [QC].[dbo].[CMMFormdata] where FormName = '"+form_name+"'")
        form_manager_list = cursor.fetchall()

        print(len(form_manager_list))

        # Nếu lấy dữ liệu ra trống
        if form_manager_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(form_manager_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for item in form_manager_list:
            position, circleNum, characteristic, link = item
            ret["data"] += [{
                "position":position,
                "circleNum":circleNum,
                "characteristic":characteristic,
                "link":link
            }]
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "form_manager").append_new_line()
        return jsonify(ret),500
    
@CMM.get('/dmc_data/<string:dmc>') 
@swag_from('./docs/CMM/dmc_data.yaml')
@token_required
def get_dmc_data(dmc):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        if dmc =="":
            dmc= "datnon"

        cursor.execute("SELECT id, actual FROM [QC].[dbo].[CMMdata] where DMC like '%"+dmc+"%' order by IDx")
        dmc_data_list = cursor.fetchall()

        print(len(dmc_data_list))

        # Nếu lấy dữ liệu ra trống
        if dmc_data_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(dmc_data_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)
    

        for idx, item in enumerate(dmc_data_list):
            id, actual = item
            ret["data"] += [{
                "id":id,
                "actual":actual
            }]
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "dmc_data").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/sixpack')
@swag_from('./docs/CMM/create_cmm_sixpacking.yaml')
@token_required_and_permissions
def create_cmm_sixpacking(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 14
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401
        
        # Lấy dữ liệu datetime_start, datetime_end, location
        day_start = request.json['dayStart']
        day_end = request.json['dayEnd']
        time_start = request.json['timeStart']
        time_end = request.json['timeEnd']

        datetime_start = day_start + " " + time_start
        datetime_end = day_end + " " + time_end

        # Lấy dữ liệu dmc_product, data, name
        dmc_product = request.json['dmcProduct']
        name = request.json['name']

        # url để lấy dữ liệu từ api web của A. Học
        url = 'http://192.168.8.21:5008/six_pack_v2_1'

        print("SELECT [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+name+"' and Product like '%"+dmc_product+"%' and TimeSave>'"+datetime_start+"' and TimeSave<'"+datetime_end+"'  order by TimeSave")
        cursor.execute("SELECT [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+name+"' and Product like '%"+dmc_product+"%' and TimeSave>'"+datetime_start+"' and TimeSave<'"+datetime_end+"'  order by TimeSave")
        all_records = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if all_records == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu có máy
        ret ={
                'status':True,
                'message':'Success',
                'data':[]
            }
        
        if len(all_records) == 0:
            ret['data'] = None
            return jsonify(ret)
        
        cmd = "SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+name+"' and Product like '%"+dmc_product+"%' and TimeSave>'"+datetime_start+"' and TimeSave<'"+datetime_end+"'  order by TimeSave"
        data2 = pd.read_sql("SELECT  [actual],[nominal],[uppertol],[lowertol] FROM [QC].[dbo].[CMMdata] where id = '"+name+"' and Product like '%"+dmc_product+"%' and TimeSave>'"+datetime_start+"' and TimeSave<'"+datetime_end+"'  order by TimeSave",conn)
        
        lst = []
        ust = []
        
        for x in data2["lowertol"]:
            try:
                lst.append(float(x))
            except:
                pass

        for x in data2["uppertol"]:
            try:
                ust.append(float(x))
            except:
                pass

        if len(lst) == 0:
            lsl = None
        else:
            lsl = sum(lst) / len(lst)+float(data2.loc[0,"nominal"])

        if len(ust) == 0:
            usl = None
        else:
            usl = sum(ust) / len(ust)+float(data2.loc[0,"nominal"])
        \
        data = {
        "LSL":  lsl,
        "USL": usl,
        "data": [float(x) for x in data2["actual"].tolist()],
        "name": name
        }

        print(data)
            
        if len(data['data']) > 0:
            Req = requests.post(url=url,json=data)
            ret["data"] = Req.text
        else:
            ret["data"] = []
            return jsonify(ret), 400

        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "cmm_sixpack").append_new_line()
        return jsonify(ret),500
    
# Manage Form
@CMM.post('/form_manager/duplicate') 
@swag_from('./docs/CMM/duplicate_form.yaml')
@token_required_and_permissions
def duplicate_form(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 10
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 10
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401
        
        # Lấy dữ liệu oldForm, newForm, fileName
        oldForm = request.json['oldForm']
        newForm = request.json['newForm']
        fileName = request.json['fileName']

        # Tạo mới một folder mới để chứa dữ liệu duplicate
        os.makedirs('C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + newForm)
        shutil.copy('C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + oldForm + '/' + fileName, 'C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + newForm + '/' + fileName)

        # Check xem newForm có tồn tại hay chưa, nếu rồi thì phải về lỗi
        cursor.execute("select id from CMMFormdata where FormName = '" + newForm + "'")
        check_exist_newform = cursor.fetchall()

        if len(check_exist_newform) > 0:
            ret = {
                'status':False,
                'message':'Already exist this form name!'
            }
            return jsonify(ret),400

        # Khai báo dict để trả về khi chưa tồn tại newForm
        ret = {
                'status':True,
                'message':'Duplicated form successfully!',
                'data':[]
            }
       
        # Lưu dữ liệu 
        print("INSERT INTO [CMMFormData] (FormName, FileName, CircleNum, Characteristic, Link, Position) SELECT '"+newForm+"' AS FormName,FileName,CircleNum,Characteristic,Link,Position FROM [CMMFormData] WHERE FormName = '"+oldForm+"'")
        cursor.execute("INSERT INTO [CMMFormData] (FormName, FileName, CircleNum, Characteristic, Link, Position) SELECT '"+newForm+"' AS FormName,FileName,CircleNum,Characteristic,Link,Position FROM [CMMFormData] WHERE FormName = '"+oldForm+"'")
        cursor.commit()

        # Trả về thông tin các form_name sau khi đã cập nhật
        cursor.execute("select distinct(FormName), FileName from CMMFormdata group by FormName, FileName")
        all_data = cursor.fetchall()

        # Thêm dữ liệu vô ret
        for item in all_data:
            form_name, file_name = item
            ret["data"].append({
                'formName': form_name,
                'fileName': file_name
            })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message':'Duplicated form fail!',
            'error':str(e)
        }
        Systemp_log1(traceback.format_exc(), "duplicate_form").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/form_manager/delete') 
@swag_from('./docs/CMM/delete_form.yaml')
@token_required_and_permissions
def delete_form(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 10
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Lấy dữ liệu user_name, password, form_name
        user_name = request.json['userName']
        password = request.json['password']
        form_name = request.json['formName']

        # Check xem có đúng user_name và password không?
        print("SELECT Top(1) * FROM [Auto].[dbo].[Web_User] where username = '"+user_name+"' and password = '"+password+"'")
        cursor.execute("SELECT Top(1) * FROM [Auto].[dbo].[Web_User] where username = '"+user_name+"' and password = '"+password+"'")
        confirm_user = cursor.fetchall()

        # Khai báo ret
        ret = {}

        if len(confirm_user):
            # Khai báo dict để trả về
            ret = {
                    'status':True,
                    'message':'Deleted form successfully!',
                    'data':[]
                }
        
            # Xóa thư mục từ máy chủ
            shutil.rmtree('C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + form_name)
            
            # Xóa form từ database
            cursor.execute("delete from cmmformdata where formname = '"+ form_name +"'")
            cursor.commit()

            # Check xem form có được xóa thành công hay chưa, nếu đã xóa thành công thì trả về list các formName và Filename sau khi đã xóa
            if len(pd.read_sql("select id from CMMFormdata where FormName = '" + form_name + "'", conn)) == 0:            
                cursor.execute("select distinct(FormName), FileName from CMMFormdata group by FormName, FileName")
                all_data = cursor.fetchall()

                # Thêm dữ liệu vô ret
                for item in all_data:
                    form_name, file_name = item
                    ret["data"].append({
                        'formName': form_name,
                        'fileName': file_name
                    })
        else:
            ret = {
                'status':False,
                'message':'Verify user fail'
            }
            return jsonify(ret),400
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message':'Delete form fail',
            'error':str(e)
        }
        Systemp_log1(traceback.format_exc(), "delete_form").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/form_manager/edit') 
@swag_from('./docs/CMM/edit_form.yaml')
@token_required_and_permissions
def edit_form(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 10
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Lấy dữ liệu oldForm, newForm, fileName
        form_name = request.json['formName']
        file_name = request.json['fileName']
        updated_data = request.json['updatedData']
        
        # Xóa form_name đó trong bảng CMMFormData
        cursor.execute("delete from cmmformdata where formname = '"+ form_name +"'")
        cursor.commit()

        # Thêm dữ liệu mới updated vào lại bảng CMMFormData
        count = 0
        for info in updated_data:
            cursor.execute( 
                '''INSERT INTO CMMFormData VALUES (?,?,?,?,?,?) ''',
                    form_name, 
                    file_name,
                    info["circleNum"], 
                    info["characteristic"],    
                    info["link"],
                    count,
            )
            count +=1

        # Chạy các câu lệnh phía trên
        cursor.commit()

        # Khai báo dict để trả về
        ret = {
                'status':True,
                'message':'Updated form successfully!'
            }
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message':'Update form fail',
            'error':str(e)
        }
        Systemp_log1(traceback.format_exc(), "edit_form").append_new_line()
        return jsonify(ret),500
    
# Sixpack Form
@CMM.get('/form_sixpack/<productCode>')
@swag_from('./docs/CMM/form_sixpack_detail.yaml')
@token_required
def get_form_sixpack_products(productCode):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()
        
        # Xóa form_name đó trong bảng CMMFormData
        cursor.execute("select distinct(a.id), b.cmm from [QC].[dbo].[CMMdata] a left join [QC].[dbo].[CMMsixpackform] b on a.id = b.cmm where a.Product like '%"+productCode+"%' order by a.id")
        all_sixpack_forms = cursor.fetchall()

        # Nếu lấy dữ liệu không thành công
        if all_sixpack_forms == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        # Nếu dữ liệu trống
        if len(all_sixpack_forms) == 0:
            ret['data'] = None                  
            return jsonify(ret)
        
        # Lưu dữ liệu vô ret
        for item in all_sixpack_forms:
            cmm, name = item
            ret["data"].append({
                "CMM": cmm,
                "Name": name
            })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "form_sixpack").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/form_sixpack/save') 
@swag_from('./docs/CMM/save_form_sixpack.yaml')
@token_required
def save_form_sixpack():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
        cursor = conn.cursor()

        # Lấy dữ liệu oldForm, newForm, fileName
        product_dmc = request.json['productDmc']
        updated_data = request.json['updatedData']
        
        # Xóa form_name đó trong bảng CMMFormData
        cursor.execute("delete FROM [QC].[dbo].[CMMsixpackform] where product = '"+product_dmc+"'")
        cursor.commit()

        # Thêm dữ liệu mới updated vào lại bảng CMMFormData
        for info in updated_data:
            cursor.execute("insert into [QC].[dbo].[CMMsixpackform] values (?,?,?)",
                        product_dmc,
                        info["CMM"],
                        info["Name"])

        # Chạy các câu lệnh phía trên
        cursor.commit()

        # Khai báo dict để trả về
        ret = {
                'status':True,
                'message':'Update form successfully.'
            }
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message':'Update form fail',
            'error':str(e)
        }
        Systemp_log1(traceback.format_exc(), "edit_form").append_new_line()
        return jsonify(ret),500
    
# CREATE FORM
@CMM.post('/upload_form/<string:ref>') 
@swag_from('./docs/CMM/upload_cmm_form.yaml')
@token_required
def upload_form(ref):
    try:
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234; Database=KnifeCNCSystem; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Kiểm tra xem 'data' có nằm trong list files không
        print(request.files)
        if 'data' not in request.files:
            ret = {
                'status':False,
                'message':'No file part'
            }
            return jsonify(ret),400

        # Lấy file storage object
        uploaded_file = request.files['data']

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                "data": []
            }
        
        # Xử lý file Excel tại đây (ví dụ: lưu file vào thư mục, đọc dữ liệu từ file, ...)
        # Dưới đây là ví dụ lưu file vào thư mục uploads
        file_path = 'C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + uploaded_file.filename
        uploaded_file.save(file_path)
        
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Xử lý dữ liệu và trả về kết quả
        row_number = 24  # Hàng bắt đầu từ 23 (C23, D23)
        ref_flag = False
        if ref != "No Reference":
            conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)  
            olddata = pd.read_sql("select * from cmmformdata where formname = '"+ref+"'",conn)
            ref_flag = True

        while True:
            # Lấy các thông số từ file_uploaded
            num = sheet.cell(row=row_number, column=4).value
            charr = sheet.cell(row=row_number, column=5).value
            norminal = sheet.cell(row=row_number, column=7).value
            lower = sheet.cell(row=row_number, column=8).value
            upper = sheet.cell(row=row_number, column=9).value

            # Khi nào num là None thì break vòng lặp -> Ngưng lấy dữ liệu
            if num is None:
                break
            link = ''
            
            # Kiểm tra xem có reference tới mẫu khác không
            if ref_flag:
                for i in range(len(olddata)):
                    if str(olddata.loc[i,"CircleNum"]) == str(num):
                        link = str(olddata.loc[i,"Link"])
                        break

            ret['data'].append({"Num":num,
                                    "Char":charr,
                                    "Norminal":norminal,
                                    "Lower":lower,
                                    "Upper":upper,
                                    "Link":link
                                    })
            row_number += 1

        return ret
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "upload_cmm_form").append_new_line()
        return jsonify(ret),500
    
@CMM.post('/form_manager/create') 
@swag_from('./docs/CMM/save_new_form.yaml')
@token_required_and_permissions
def save_new_form(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)         
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 9
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        # Kiểm tra xem 'data' có nằm trong list files không
        if 'data' not in request.files:
            ret = {
                'status':False,
                'message':'No file part!'
            }
            return jsonify(ret),400

        # Lấy file storage object
        uploaded_file = request.files['data']

        # Lấy dữ liệu oldForm, newForm, fileName
        form_name = request.form['formName']
        file_name = request.form['fileName']
        updated_data = json.loads(request.form['updatedData'])

        # Kiểm tra xem tên folder có bị trung không
        cursor.execute("SELECT FormName as count From CMMFormData where FormName = '"+form_name+"'")
        check_form_name = cursor.fetchall()

        if len(check_form_name) > 0:
            ret = {
                'status':False,
                'message':'Form name already exists!'
            }
            return jsonify(ret),400
        
        # Nếu tên folder không trùng
        ret = {
                'status':True,
                'message':'Create form name successfully!',
            }

        os.makedirs('C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + form_name)
        file_path = 'C:/inetpub/wwwroot/WEB_DATA_API/cmm_forms/' + form_name + '/' + file_name
        uploaded_file.save(file_path)
        
        print(updated_data)
        count = 0
        for info in updated_data:
            print(info)
            cursor.execute( 
                '''INSERT INTO CMMFormData VALUES (?,?,?,?,?,?) ''',
                    form_name, 
                    file_name,
                    info["circleNumber"], 
                    info["characteristic"],    
                    info["link"],
                    count,
            )
            count +=1

        # Thực hiện toàn bộ các 
        cursor.commit()
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message':'Create form fail!',
            'error':str(e)
        }
        Systemp_log1(traceback.format_exc(), "save_new_form").append_new_line()
        return jsonify(ret),500
    
# Lấy danh sách tất cả reference
@CMM.get('/references') 
@swag_from('./docs/CMM/get_references.yaml')
@token_required
def get_references():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("SELECT DISTINCT(FormName) FROM [QC].[dbo].[CMMFormData]")
        reference_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if reference_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(reference_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)

        for item in reference_list:
            file_name = item
            ret["data"].append(file_name[0])
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "references").append_new_line()
        return jsonify(ret),500
    
# Lấy danh sách tất cả reference
@CMM.get('/links/<string:reference>') 
@swag_from('./docs/CMM/get_links.yaml')
@token_required
def get_links(reference):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        cursor.execute("SELECT Characteristic, Link FROM [QC].[dbo].[CMMFormData] where FormName = '"+ reference +"'")
        link_list = cursor.fetchall()

        # Nếu lấy dữ liệu ra trống
        if link_list == None:
            ret = {
                'status':False,
                'message':'Not Exist Data'
            }
            return jsonify(ret),400

        # Nếu all_records được truy vấn thành công
        ret = {
                'status':True,
                'message':'Success',
                'data': []
            }
        
        if len(link_list) == 0:
            ret['data'] = None                  
            return jsonify(ret)

        for item in link_list:
            characteristic, link = item
            ret["data"].append({
                'characteristic': characteristic,
                'link': link
            })
        
        return jsonify(ret)
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "links").append_new_line()
        return jsonify(ret),500
    
# @CMM.post('/form_manager/edit') 
# @swag_from('./docs/CMM/edit_form.yaml')
# @token_required
# def edit_form():
#     try:
#         # Tạo cusor để kết nối với database
#         conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)        
#         cursor = conn.cursor()

#         # Lấy dữ liệu oldForm, newForm, fileName
#         form_name = request.json['formName']
#         file_name = request.json['fileName']
#         updated_data = request.json['updatedData']
        
#         # Xóa form_name đó trong bảng CMMFormData
#         cursor.execute("delete from cmmformdata where formname = '"+ form_name +"'")
#         cursor.commit()

#         # Thêm dữ liệu mới updated vào lại bảng CMMFormData
#         count = 0
#         for info in updated_data:
#             cursor.execute( 
#                 '''INSERT INTO CMMFormData VALUES (?,?,?,?,?,?) ''',
#                     form_name, 
#                     file_name,
#                     info["circleNum"], 
#                     info["characteristic"],    
#                     info["link"],
#                     count,
#             )
#             count +=1

#         # Chạy các câu lệnh phía trên
#         cursor.commit()

#         # Khai báo dict để trả về
#         ret = {
#                 'status':True,
#                 'message':'Updated form successfully!'
#             }
        
#         return jsonify(ret)
#     except Exception as e:
#         ret = {
#             'status':False,
#             'message':'Update form fail',
#             'error':str(e)
#         }
#         Systemp_log1(traceback.format_exc(), "edit_form").append_new_line()
#         return jsonify(ret),500
    
@CMM.post('/update_purpose')
@swag_from('./docs/CMM/update_purpose_data.yaml')
@token_required_and_permissions
def update_purpose_data(role, permissions):
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)
        cursor = conn.cursor()

        # Check xem có phải Manager không?
        if role < 100:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),401

        # Check xem có permission không?
        id_permission = 11
        binary_number = bin(int(permissions, 16))[2:]

        cursor.execute("select count(*) from [QC].[dbo].[Permission_Name]")
        num_permissions = cursor.fetchone()

        binary_number = binary_number.zfill(num_permissions[0])
        permission_value = binary_number[-id_permission]

        if not int(permission_value):
            ret = {
                'status':False,
                'message':'Sorry, your role does not have this permission!'
            }
            return jsonify(ret),401

        if role < 2:
            ret = {
                'status':False,
                'message':'Sorry, permission denied!'
            }
            return jsonify(ret),400

        id = request.json['id']
        purpose = request.json['purpose']

        # check if cmm records is not exist
        cursor.execute("select * from [QC].[dbo].[CMMdata] where IDx = '" + str(id) + "'")
        count_record_dups = cursor.fetchall()

        if count_record_dups == None or len(count_record_dups) == 0:
            return jsonify({
                        'status':False,
                        'message':'This data does not exist!'
                    })  

        print("update [QC].[dbo].[CMMdata] set CodePurpose = '" + purpose + "' where IDx = '" + str(id) + "'")
        cursor.execute("update [QC].[dbo].[CMMdata] set CodePurpose = '" + purpose + "' where IDx = '" + str(id) + "'")
        
        conn.commit()

        # if insert data to Vending table successfully
        ret = {
            'status':True,
            'message':'Update data successfully!',
            'data': {'id': id, 'purpose': purpose}
        }
        return jsonify(ret), HTTP_201_CREATED
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(str(e), "update_purpose_data").append_new_line()
        return jsonify(ret), 500

import pyodbc
import openpyxl
from openpyxl import load_workbook,styles
from openpyxl.styles import Alignment
from openpyxl.styles import Color, Fill, Border, Side
from openpyxl.styles import Font
import io
from flask import *

@CMM.post('/export_form')
@swag_from('./docs/CMM/export_CMM_form.yaml')
@token_required
def export_CMM_form():
    try:
        # Tạo cusor để kết nối với database
        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)

        # Lấy dữ liệu formName, dmc
        formname = request.json["formName"]
        dmclist1 = request.json["dmc"]   

        dmclist = []
        for item in dmclist1:
            if item not in dmclist:
                dmclist.append(item)

        conn = pyodbc.connect('Driver={SQL Server}; Server=192.168.8.21; uid=sa; pwd=1234;Database=QC; Trusted_Connection=No;',timeout=1)   
        filename = pd.read_sql("select top 1 filename from cmmformdata where Formname ='"+formname+"'",conn).loc[0,"filename"]
        wb_sample = load_workbook('cmmform/'+formname+'/'+filename)
        sheet=wb_sample.active
        my_red = openpyxl.styles.colors.Color(rgb='00ff5555')
        my_gray = openpyxl.styles.colors.Color(rgb='00d9d9d9')
        daycl = openpyxl.styles.colors.Color(rgb='0055ff55')
        nightcl = openpyxl.styles.colors.Color(rgb='00FFA500')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_fill2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_gray)
        dayfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=daycl)
        nightfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=nightcl)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border1 = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        
        if len(dmclist) > 3:
            try:
                sheet.unmerge_cells(start_row=21,end_row=22,start_column=19,end_column=19)
            except:
                pass
            sheet.insert_cols(17,len(dmclist)-3)
            for i in range(len(dmclist)-3):
                
                for row_num in range(23, sheet.max_row + 1):
                    sheet.cell(row_num,17+i).border = border1
                    sheet.cell(row_num,17+i).alignment = Alignment(horizontal='center',vertical='center')
                    sheet.cell(row_num,17+i).font = Font(name="Arial")
                for row in range(2,10):
                    sheet.cell(row,17+i).border = Border(left=Side(style='thin'), right=Side(style='dashed'), top=Side(style='thin'), bottom=Side(style='thin'))
                for row in range(11,21):
                    sheet.cell(row,17+i).border = Border(top=Side(style='dashed'))
                sheet.cell(10,17+i).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                sheet.cell(21,17+i).border = border
                sheet.cell(22,17+i).border = border
                sheet.cell(23,17+i).border = border
                sheet.cell(21,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(22,17+i).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(23,17+i).alignment = Alignment(horizontal='center',vertical='center')
                sheet.cell(21,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).font = Font(name="Arial")
                sheet.cell(23,17+i).font = Font(name="Arial")
                sheet.cell(22,17+i).fill = my_fill2
    
        for idmc in range(len(dmclist)):
            try:
                idlist = pd.read_sql("select link as idx from cmmformdata where formname = '"+formname+"' order by Position",conn)
                data = pd.read_sql("SELECT [DMC],[Line],[TimeSave],[id],[actual],[Result],[CodePurpose] FROM [QC].[dbo].[CMMdata] where dmc = N'"+dmclist[idmc]+"'",conn)
                print(idmc,len(dmclist))
                if data.loc[0,"TimeSave"].hour < 8 or data.loc[0,"TimeSave"].hour > 19:
                    sheet.cell(21,idmc+16).fill = nightfill
                else:
                    sheet.cell(21,idmc+16).fill = dayfill
                sheet.cell(21,idmc+16).value = str(data.loc[0,"TimeSave"])[:-7]
                sheet.cell(22,idmc+16).value = data.loc[0,"Line"] +' '+ dmclist[idmc]
                print(data.loc[0,"Line"] +' '+ dmclist[idmc])
                sheet.cell(23,idmc+16).value = data.loc[0,"CodePurpose"]
                sheet.cell(23,idmc+16).border = border
                data.set_index("id", inplace=True)
                for i in range(len(idlist)):
                    try:
                        sheet.cell(24+i,idmc+16).value = round(float(data.loc[idlist.loc[i,"idx"],"actual"]),3)
                        sheet.cell(24+i,idmc+16).number_format = '0.000'
                        if "No Tol" not in idlist.loc[i,"idx"]:                            
                            if data.loc[idlist.loc[i,"idx"],"Result"].strip() == "":
                                if data.loc[idlist.loc[i,"idx"].replace(".XA",'').replace(".YA",'').replace(".ZA",'').replace(".R",'').replace(".PH",''),"Result"].strip() != "OK":
                                    sheet.cell(24+i,idmc+16).fill = my_fill
                            elif data.loc[idlist.loc[i,"idx"],"Result"].strip() != "OK":
                                sheet.cell(24+i,idmc+16).fill = my_fill
                    except Exception as e:
                        sheet.cell(24+i,idmc+16).fill = my_fill

            except:
                print(dmclist[idmc])
        output = io.BytesIO()

        # add headers
        wb_sample.save(output)
        # writer.save()
        output.seek(0)

        return Response(output, mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=CMM_Data.xls"})
    
    except Exception as e:
        ret = {
            'status':False,
            'message': str(e)
        }
        Systemp_log1(traceback.format_exc(), "export_CMM_form").append_new_line()
        return jsonify(ret),500